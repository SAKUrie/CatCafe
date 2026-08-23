"""策划表与项目配置的同步检查。只读，不改任何东西。

用法：python Tools/CatCafeConfig/check_design_sync.py
输出：Docs/design_sync_report.txt

比两件事：
  A. 上一版策划表 -> 最终版：策划改了什么（增删条目、稀有度、效果）
  B. 最终版 -> 项目 Pieces/Buffs：稀有度、效果文案、图标还差什么

策划表更新后重跑一次即可；两份表的路径见下面的 V3 / FINAL 常量。
名称匹配用的是精确同名——策划改了名字会显示成"两边各自独有"，
这种要人工看一眼（历史上出现过 英短 -> 英短猫 这类加后缀的改名）。
"""
import io
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
V3 = Path(r'C:\Users\tengfei\Downloads\猫咖经营_店内对象设计_V3(1).xlsx')
FINAL = Path(r'C:\Users\tengfei\Downloads\猫咖经营_店内对象与道具-最终版.xlsx')
CFG = json.load(io.open(ROOT / 'Assets' / 'Resources' / 'GameData' / 'cat_cafe_config.json',
                        encoding='utf-8'))
RES = ROOT / 'Assets' / 'Resources' / 'CatCafe'


def norm(s):
    return re.sub(r'[\s　]', '', str(s or '').replace('\\n', '').replace('\n', ''))


def read(path, sheet, header_row):
    ws = openpyxl.load_workbook(path, data_only=True)[sheet]
    hdr = [str(c.value or '').strip() for c in ws[header_row]]
    idx = {n: i for i, n in enumerate(hdr) if n}
    out = {}
    for r in range(header_row + 1, ws.max_row + 1):
        v = [c.value for c in ws[r]]
        if '名称' not in idx or idx['名称'] >= len(v) or not v[idx['名称']]:
            continue
        get = lambda k: (str(v[idx[k]]).strip()
                         if k in idx and idx[k] < len(v) and v[idx[k]] is not None else '')
        out[str(v[idx['名称']]).strip()] = {
            'no': get('编号'), 'school': get('所属流派'),
            'rarity': get('稀有度'), 'gold': get('金币'), 'effect': get('效果'),
        }
    return out


def main() -> int:
    out = io.open(ROOT / 'Docs' / 'design_sync_report.txt', 'w', encoding='utf-8')

    for label, sheet_v3, hr_v3, sheet_f, hr_f in [
            ('店内对象', '店内对象', 4, '店内对象', 4),
            ('道具', '道具设计', 6, '道具', 6)]:
        a = read(V3, sheet_v3, hr_v3)
        b = read(FINAL, sheet_f, hr_f)
        out.write('=' * 72 + '\nA. %s：V3 → 最终版\n' % label + '=' * 72 + '\n')
        out.write('V3 %d 条 → 最终版 %d 条\n' % (len(a), len(b)))
        removed = sorted(set(a) - set(b))
        added = sorted(set(b) - set(a))
        out.write('  删除 %d：%s\n' % (len(removed), '、'.join(removed) if removed else '无'))
        out.write('  新增 %d：%s\n' % (len(added), '、'.join(added) if added else '无'))
        rar = [(n, a[n]['rarity'], b[n]['rarity']) for n in sorted(set(a) & set(b))
               if a[n]['rarity'] != b[n]['rarity']]
        eff = [(n, a[n]['effect'], b[n]['effect']) for n in sorted(set(a) & set(b))
               if norm(a[n]['effect']) != norm(b[n]['effect'])]
        out.write('  稀有度变动 %d：\n' % len(rar))
        for n, x, y in rar:
            out.write('     %-14s %s → %s\n' % (n, x, y))
        out.write('  效果变动 %d：\n' % len(eff))
        for n, x, y in eff:
            out.write('     %s\n       V3 : %s\n       最终: %s\n' % (n, x, y))
        out.write('\n')

    # 最终版 → 项目
    order = [r['key'] for r in sorted(CFG['rarities'], key=lambda x: x['index'])]
    RMAP = dict(zip(['普通', '稀有', '史诗', '传奇'], order))
    assets = {p.stem for p in RES.rglob('*.png')}

    for label, sheet, hr, proj in [('店内对象', '店内对象', 4, CFG['elements']),
                                   ('道具', '道具', 6, CFG['items'])]:
        f = read(FINAL, sheet, hr)
        pn = {}
        for p in proj:
            pn.setdefault(str(p.get('name', '')).strip(), p)
        both = sorted(set(f) & set(pn))
        out.write('=' * 72 + '\nB. %s：最终版 ↔ 项目\n' % label + '=' * 72 + '\n')
        out.write('最终版 %d ｜项目 %d ｜同名 %d ｜项目独有 %d ｜待接入 %d\n\n'
                  % (len(f), len(proj), len(both), len(set(pn) - set(f)), len(set(f) - set(pn))))
        rbad = [(n, f[n]['rarity'], RMAP.get(f[n]['rarity']), pn[n].get('rarity'))
                for n in both if RMAP.get(f[n]['rarity']) != pn[n].get('rarity')]
        ebad = [(n, f[n]['effect'], pn[n].get('rule_text', ''))
                for n in both if norm(f[n]['effect']) != norm(pn[n].get('rule_text', ''))]
        out.write('  稀有度不一致 %d / %d\n' % (len(rbad), len(both)))
        for n, a_, want, got in rbad:
            out.write('     %-14s 最终=%s(%s)  项目=%s\n' % (n, a_, want, got))
        out.write('  效果不一致 %d / %d\n' % (len(ebad), len(both)))
        for n, x, y in ebad:
            out.write('     %s\n       最终: %s\n       项目: %s\n' % (n, x, str(y).replace('\n', ' / ')))
        miss = [p.get('name') for p in proj
                if not str(p.get('asset', '') or '').strip()
                or str(p.get('asset')).split('/')[-1] not in assets]
        out.write('  图标缺失 %d：%s\n\n' % (len(miss), '、'.join(miss) if miss else '无'))

    out.close()
    print('done')
    return 0


if __name__ == '__main__':
    sys.exit(main())
