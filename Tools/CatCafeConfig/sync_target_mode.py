#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""把目标模式中已经同效果的条目同步名称/稀有度，并刷新闭包状态。

脚本不创造机制、不启用条目；只有目标效果与正式 rule_text 已一致时才对齐身份。
配合 audit_target_mode.py 使用，避免人工把不完整内容标成“已接入”。
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from xlsx_patch import Workbook  # noqa: E402


ROOT = Path(__file__).resolve().parents[2]
BOOK = ROOT / "GameDesign" / "CatCafeGameConfig.xlsx"
RARITY = {
    "普通": "common", "稀有": "uncommon", "史诗": "rare",
    "传奇": "special", "特殊": "special",
}


def norm(value) -> str:
    return re.sub(r"[\s　，。；、：:（）()]", "", str(value or ""))


def truthy(value) -> bool:
    return isinstance(value, bool) and value or str(value or "").strip().lower() in {
        "true", "1", "yes", "y"
    }


def source_rows(book, sheet_name: str) -> dict:
    sheet = book[sheet_name]
    fields = [cell.value for cell in sheet[3]]
    return {
        row[fields.index("key")].value: {
            fields[i]: row[i].value for i in range(len(fields))
        }
        for row in sheet.iter_rows(min_row=5)
        if row[fields.index("key")].value
    }


def main() -> int:
    source_book = openpyxl.load_workbook(BOOK, data_only=True)
    pieces = source_rows(source_book, "Pieces")
    items = source_rows(source_book, "Buffs")
    status = source_book["V3接入状态"]
    headers = [cell.value for cell in status[2]]
    rule_sheet = source_book["Rules"]
    rule_headers = [cell.value for cell in rule_sheet[3]]
    enabled_rules = []
    for row in rule_sheet.iter_rows(min_row=5):
        values = {rule_headers[i]: row[i].value for i in range(len(rule_headers))}
        if truthy(values.get("enabled")):
            enabled_rules.append(values)

    identity_edits = []
    status_edits = []
    for row_number, row in enumerate(status.iter_rows(min_row=3), start=3):
        category = row[headers.index("类别")].value
        key = row[headers.index("配置键")].value
        source = pieces.get(key) if category == "店内对象" else items.get(key) if category == "长期道具" else None
        if source is None:
            continue
        text_exact = norm(row[headers.index("原始效果")].value) == norm(source.get("rule_text"))
        target_rarity = RARITY.get(str(row[headers.index("稀有度")].value or ""))
        if text_exact and target_rarity:
            identity_edits.append((category, key, row[headers.index("名称")].value,
                                   target_rarity, bool(source.get("pool_rarity"))))
            source["name"] = row[headers.index("名称")].value
            source["rarity"] = target_rarity

        exact = text_exact and target_rarity == str(source.get("rarity") or "") and \
            norm(row[headers.index("名称")].value) == norm(source.get("name"))
        enabled = truthy(source.get("enabled"))
        target_text = str(row[headers.index("原始效果")].value or "")
        if exact and not enabled and target_text.startswith("未开放："):
            state = "目标保留禁用"
            note = "目标模式明确标注未开放；保持禁用即为正确运行状态。"
        elif exact and enabled:
            state = "已接入"
            note = "目标模式已闭包：配置、运行时与自动回归均已接通。"
        elif not enabled:
            state = "暂缓，未启用"
            note = "目标模式尚未闭包；保持禁用，避免不完整内容进入奖励池。"
        else:
            state = "待目标模式闭包"
            note = "目标效果与当前正式配置不一致，禁止视为闭包。"
        if row[headers.index("配置状态")].value != state or row[headers.index("接入说明")].value != note:
            status_edits.append((row_number, state, note))
    source_book.close()

    workbook = Workbook(str(BOOK))
    for category, key, name, rarity, has_pool in identity_edits:
        sheet = "Pieces" if category == "店内对象" else "Buffs"
        workbook.set_cell(sheet, key, "name", name)
        workbook.set_cell(sheet, key, "rarity", rarity)
        if sheet == "Pieces" and has_pool:
            workbook.set_cell(sheet, key, "pool_rarity", rarity)
    for row_number, state, note in status_edits:
        workbook.set_cell_ref("V3接入状态", row_number, "H", state)
        workbook.set_cell_ref("V3接入状态", row_number, "J", note)
    operations = sorted({str(row.get("operation") or "") for row in enabled_rules if row.get("operation")})
    scopes = sorted({str(row.get(field) or "") for row in enabled_rules
                     for field in ("primary_scope", "secondary_scope") if row.get(field)})
    workbook.set_cell_ref("说明", 9, "B", " / ".join(operations))
    workbook.set_cell_ref("说明", 10, "B", " / ".join(scopes))
    workbook.save(backup=False)
    print(f"identity_rows={len(identity_edits)}, status_rows={len(status_edits)}, edits={workbook.edits}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
